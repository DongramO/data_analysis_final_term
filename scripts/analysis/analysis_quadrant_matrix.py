# -*- coding: utf-8 -*-
"""
4분면 매트릭스: X축=매출QoQ복합점수 / Y축=이동성장세점수
각 사분면 내 상위 15개 행정동 이름 표시
출력: image/quadrant_matrix.png
      data/quadrant_matrix.xlsx
"""
import warnings
import numpy as np
import pandas as pd
import matplotlib.pyplot as plt
import matplotlib.patches as mpatches
from scipy.stats import zscore as sp_zscore, linregress
from pathlib import Path
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule

warnings.filterwarnings('ignore')
plt.rcParams['font.family'] = 'Malgun Gothic'
plt.rcParams['axes.unicode_minus'] = False

ROOT     = Path('.')
RF_PATH  = ROOT / 'data/main_data/분기별_원본피처.csv'
MIG_PATH = ROOT / 'data/main_data/분기별_주말유입인구.csv'
MAP_PATH = ROOT / 'data/main_data/코드매핑_유입인구_매출.csv'

def clip_z(s):
    lo, hi = s.quantile(0.005), s.quantile(0.995)
    return pd.Series(sp_zscore(s.clip(lo, hi), ddof=0), index=s.index)

# ════════════════════════════════════════════════════════════
# ① 매출 QoQ 복합점수
# ════════════════════════════════════════════════════════════
df = pd.read_csv(RF_PATH, encoding='utf-8-sig')

ALL_Q_S   = ['2022Q4','2023Q1','2023Q2','2023Q3','2023Q4',
             '2024Q1','2024Q2','2024Q3','2024Q4']
ANA_Q_S   = ALL_Q_S[1:]

sub_s = df[df['연도분기'].isin(ALL_Q_S)]
qrt = (sub_s.groupby(['행정동','연도분기'])[['연령대_20_매출_금액','총_매출금액']]
            .sum())

piv_s20  = qrt['연령대_20_매출_금액'].unstack('연도분기')
piv_stot = qrt['총_매출금액'].unstack('연도분기')

piv_s20  = piv_s20.dropna(subset=ALL_Q_S)
piv_stot = piv_stot.reindex(piv_s20.index).dropna(subset=ALL_Q_S)
common_s = piv_s20.index.intersection(piv_stot.index)
piv_s20  = piv_s20.loc[common_s]
piv_stot = piv_stot.loc[common_s]

z_s규모_list, z_s비중_list, z_s성장률_list = [], [], []
for i, q in enumerate(ANA_Q_S):
    prev_q = ALL_Q_S[i]
    규모   = piv_s20[q] / 1e8
    비중   = piv_s20[q] / piv_stot[q] * 100
    g      = (piv_s20[q] - piv_s20[prev_q]) / piv_s20[prev_q].abs() * 100
    z_s규모_list.append(pd.Series(sp_zscore(규모, ddof=0), index=piv_s20.index))
    z_s비중_list.append(pd.Series(sp_zscore(비중, ddof=0), index=piv_s20.index))
    z_s성장률_list.append(clip_z(g))

sales_score = pd.Series(
    (pd.concat(z_s규모_list, axis=1).mean(axis=1)
     + pd.concat(z_s비중_list, axis=1).mean(axis=1)
     + pd.concat(z_s성장률_list, axis=1).mean(axis=1)) / 3,
    name='매출QoQ복합점수'
)

# ════════════════════════════════════════════════════════════
# ② 이동 성장세점수
# ════════════════════════════════════════════════════════════
_map = pd.read_csv(MAP_PATH, encoding='utf-8-sig')
_map.columns = ['mig_code','raw_code','gu','dong_map']
code2dong = (
    pd.read_csv(RF_PATH, encoding='utf-8-sig', usecols=['행정동_코드','행정동'])
    .drop_duplicates()
    .rename(columns={'행정동_코드':'raw_code','행정동':'dong'})
    .set_index('raw_code')['dong'].to_dict()
)
mig = pd.read_csv(MIG_PATH, encoding='utf-8-sig')
mig.columns = ['mig_code','yr','q','yrq','age','HE','WE','EE','mig_total']
mig['raw_code'] = mig['mig_code'].map(_map.set_index('mig_code')['raw_code'].to_dict())
mig = mig.dropna(subset=['raw_code'])
mig['raw_code'] = mig['raw_code'].astype(int)
mig['dong'] = mig['raw_code'].map(code2dong)
mig = mig.dropna(subset=['dong'])

Q_2023   = ['2023Q1','2023Q2','2023Q3','2023Q4']
Q_2024   = ['2024Q1','2024Q2','2024Q3','2024Q4']
ANA_Q_M  = Q_2023 + Q_2024
ALL_Q_M  = ['2022Q4'] + ANA_Q_M

piv_m = (mig[mig['age'] == 20]
         .groupby(['dong','yrq'])['mig_total']
         .sum().unstack('yrq'))
piv_m = piv_m.dropna(subset=ALL_Q_M)

# 증가율
sum_23 = piv_m[Q_2023].sum(axis=1)
sum_24 = piv_m[Q_2024].sum(axis=1)
z_증가율 = clip_z((sum_24 - sum_23) / sum_23.abs() * 100)

# 기울기
X = np.arange(len(ANA_Q_M), dtype=float)
slopes = piv_m[ANA_Q_M].apply(
    lambda r: linregress(X, r.values.astype(float)).slope, axis=1)
z_기울기 = clip_z(slopes)

# 모멘텀 (QoQ 분기별 평균)
z_m_list = []
for i, q in enumerate(ANA_Q_M):
    g = (piv_m[q] - piv_m[ALL_Q_M[i]]) / piv_m[ALL_Q_M[i]].abs() * 100
    z_m_list.append(clip_z(g))
z_모멘텀 = pd.concat(z_m_list, axis=1).mean(axis=1)

mig_score = pd.Series(
    (z_증가율 + z_기울기 + z_모멘텀) / 3,
    name='이동성장세점수'
)
mig_score.index.name = '행정동'

# ════════════════════════════════════════════════════════════
# ③ 두 점수 합치기
# ════════════════════════════════════════════════════════════
merged = pd.concat([sales_score, mig_score], axis=1).dropna()
merged.index.name = '행정동'
print(f'매트릭스 행정동: {len(merged)}개')

# 사분면 분류 (기준: x=0, y=0)
def quadrant(row):
    x, y = row['매출QoQ복합점수'], row['이동성장세점수']
    if   x >= 0 and y >= 0: return 'Q1'
    elif x <  0 and y >= 0: return 'Q2'
    elif x <  0 and y <  0: return 'Q3'
    else:                   return 'Q4'

merged['사분면'] = merged.apply(quadrant, axis=1)
# 원점 거리 (사분면 내 순위 기준)
merged['거리'] = np.sqrt(merged['매출QoQ복합점수']**2 + merged['이동성장세점수']**2)

# 사분면별 TOP 15
TOP_N = 15
label_dongs = set()
quad_top = {}
for q in ['Q1','Q2','Q3','Q4']:
    sub = merged[merged['사분면'] == q].sort_values('거리', ascending=False).head(TOP_N)
    quad_top[q] = sub
    label_dongs.update(sub.index.tolist())

# ════════════════════════════════════════════════════════════
# ④ 시각화
# ════════════════════════════════════════════════════════════
QUAD_COLORS = {'Q1':'#2E75B6', 'Q2':'#2E8B57', 'Q3':'#AAAAAA', 'Q4':'#C0392B'}
QUAD_BG     = {'Q1':'#EBF3FB', 'Q2':'#E9F5EE', 'Q3':'#F5F5F5', 'Q4':'#FDEDEC'}
QUAD_LABELS = {
    'Q1': '① 매출↑  유입↑\n(쌍끌이 성장)',
    'Q2': '② 매출↓  유입↑\n(유입 선행)',
    'Q3': '③ 매출↓  유입↓\n(정체)',
    'Q4': '④ 매출↑  유입↓\n(소비 선행)',
}

fig, ax = plt.subplots(figsize=(13, 10))

# 사분면 배경
xmin, xmax = merged['매출QoQ복합점수'].min(), merged['매출QoQ복합점수'].max()
ymin, ymax = merged['이동성장세점수'].min(),   merged['이동성장세점수'].max()
pad_x = (xmax - xmin) * 0.08
pad_y = (ymax - ymin) * 0.08
xlim  = (xmin - pad_x, xmax + pad_x * 3.5)
ylim  = (ymin - pad_y, ymax + pad_y * 1.5)

from matplotlib.patches import FancyArrowPatch
import matplotlib.patches as mpatches

ax.fill_betweenx([0, ylim[1]], 0, xlim[1], color=QUAD_BG['Q1'], zorder=0)
ax.fill_betweenx([0, ylim[1]], xlim[0], 0, color=QUAD_BG['Q2'], zorder=0)
ax.fill_betweenx([ylim[0], 0], xlim[0], 0, color=QUAD_BG['Q3'], zorder=0)
ax.fill_betweenx([ylim[0], 0], 0, xlim[1], color=QUAD_BG['Q4'], zorder=0)

# 사분면 레이블 (모서리)
ax.text(xlim[1]*0.97, ylim[1]*0.95, QUAD_LABELS['Q1'],
        ha='right', va='top', fontsize=10, color=QUAD_COLORS['Q1'],
        fontweight='bold', alpha=0.7)
ax.text(xlim[0]*0.97, ylim[1]*0.95, QUAD_LABELS['Q2'],
        ha='left', va='top', fontsize=10, color=QUAD_COLORS['Q2'],
        fontweight='bold', alpha=0.7)
ax.text(xlim[0]*0.97, ylim[0]*0.95, QUAD_LABELS['Q3'],
        ha='left', va='bottom', fontsize=10, color='#888888',
        fontweight='bold', alpha=0.7)
ax.text(xlim[1]*0.97, ylim[0]*0.95, QUAD_LABELS['Q4'],
        ha='right', va='bottom', fontsize=10, color=QUAD_COLORS['Q4'],
        fontweight='bold', alpha=0.7)

# 기준선
ax.axhline(0, color='#888888', lw=1.2, zorder=1)
ax.axvline(0, color='#888888', lw=1.2, zorder=1)

# 전체 산점도 (회색 배경)
non_label = merged[~merged.index.isin(label_dongs)]
ax.scatter(non_label['매출QoQ복합점수'], non_label['이동성장세점수'],
           c='#CCCCCC', s=18, alpha=0.5, zorder=2)

# 사분면별 TOP15 강조 산점도
for q, color in QUAD_COLORS.items():
    sub = quad_top[q]
    ax.scatter(sub['매출QoQ복합점수'], sub['이동성장세점수'],
               c=color, s=55, alpha=0.85, zorder=3, edgecolors='white', linewidths=0.5)

# 라벨 표시 (간단한 오프셋 방식)
for q, color in QUAD_COLORS.items():
    sub = quad_top[q]
    for dong, row in sub.iterrows():
        x, y = row['매출QoQ복합점수'], row['이동성장세점수']
        # 사분면에 따라 라벨 방향 조정
        ha = 'left'  if x >= 0 else 'right'
        va = 'bottom' if y >= 0 else 'top'
        dx = 0.04 if x >= 0 else -0.04
        dy = 0.03 if y >= 0 else -0.03
        ax.annotate(dong,
                    xy=(x, y), xytext=(x + dx, y + dy),
                    fontsize=7.5, color=color, fontweight='bold',
                    ha=ha, va=va, zorder=4,
                    arrowprops=dict(arrowstyle='-', color=color,
                                   lw=0.5, alpha=0.6))

ax.set_xlim(xlim)
ax.set_ylim(ylim)
ax.set_xlabel('매출활성도 점수 (Stage A)', fontsize=13, fontweight='bold')
ax.set_ylabel('이동성장세 점수 (Stage B)', fontsize=13, fontweight='bold')
ax.set_title('20대 매출활성도(Stage A)  ×  주말 유입 이동성장세(Stage B)  4분면 매트릭스\n'
             '(2023Q1~2024Q4 기준  |  각 사분면 원점 거리 상위 15개 행정동 표시)',
             fontsize=13, fontweight='bold', pad=15)
ax.grid(alpha=0.15, zorder=0)
ax.spines[['top','right']].set_visible(False)

legend_handles = [mpatches.Patch(color=c, label=QUAD_LABELS[q].replace('\n',' '))
                  for q, c in QUAD_COLORS.items()]
ax.legend(handles=legend_handles, fontsize=9, loc='lower right',
          framealpha=0.9, edgecolor='#CCCCCC')

plt.tight_layout()
fig.savefig(ROOT / 'image/quadrant_matrix.png', dpi=150, bbox_inches='tight')
plt.close(fig)
print('[저장] image/quadrant_matrix.png')

# ════════════════════════════════════════════════════════════
# ⑤ 엑셀 출력
# ════════════════════════════════════════════════════════════
# 전체 순위 + 사분면 내 순위
merged = merged.sort_values('거리', ascending=False)
merged.insert(0, '전체순위', range(1, len(merged) + 1))

quad_rank = merged.groupby('사분면').cumcount() + 1
merged.insert(1, '사분면내순위', quad_rank)
merged = merged.sort_values(['사분면','사분면내순위'])

out_df = merged.reset_index()[['전체순위','사분면내순위','행정동','사분면',
                                '매출QoQ복합점수','이동성장세점수','거리']]

QUAD_ORDER = ['Q1','Q2','Q3','Q4']
QUAD_FILLS = {
    'Q1': 'D6E4F0', 'Q2': 'D5E8D4',
    'Q3': 'F0F0F0', 'Q4': 'F8D7DA'
}
QUAD_HEADER_FILLS = {
    'Q1': '2E75B6', 'Q2': '2E8B57',
    'Q3': '888888', 'Q4': 'C0392B'
}

out_path = ROOT / 'data/quadrant_matrix.xlsx'
with pd.ExcelWriter(out_path, engine='openpyxl') as writer:
    # 전체 시트
    out_df.to_excel(writer, index=False, sheet_name='전체')

    # 사분면별 시트
    for q in QUAD_ORDER:
        sub = out_df[out_df['사분면'] == q].reset_index(drop=True)
        sub.to_excel(writer, index=False, sheet_name=q)

    wb = writer.book

    def style_sheet(ws, nrow, ncol, fill_hex, header_hex, top15_only=False):
        hf    = PatternFill('solid', fgColor=header_hex)
        hfont = Font(bold=True, color='FFFFFF', size=9)
        for cell in ws[1]:
            cell.fill = hf; cell.font = hfont
            cell.alignment = Alignment(horizontal='center', vertical='center',
                                       wrap_text=True)
        ws.row_dimensions[1].height = 28

        col_widths = [8, 10, 14, 6, 16, 14, 10]
        for i, w in enumerate(col_widths[:ncol], 1):
            ws.column_dimensions[get_column_letter(i)].width = w

        top15_fill = PatternFill('solid', fgColor=fill_hex)
        alt_fill   = PatternFill('solid', fgColor='F8F8F8')
        thin = Side(style='thin', color='CCCCCC')
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        for row in range(2, nrow + 2):
            ws.row_dimensions[row].height = 15
            is_top15 = (ws.cell(row=row, column=2).value or 99) <= 15
            for ci in range(1, ncol + 1):
                cell = ws.cell(row=row, column=ci)
                cell.border = border
                cell.alignment = Alignment(horizontal='center', vertical='center')
                if is_top15:
                    cell.fill = top15_fill
                elif row % 2 == 0:
                    cell.fill = alt_fill
            ws.cell(row=row, column=3).alignment = Alignment(
                horizontal='left', vertical='center')
            # 숫자 포맷
            for ci, col_name in enumerate(['전체순위','사분면내순위','행정동','사분면',
                                            '매출QoQ복합점수','이동성장세점수','거리'], 1):
                cell = ws.cell(row=row, column=ci)
                if col_name in ['매출QoQ복합점수','이동성장세점수','거리']:
                    cell.number_format = '0.000'
            # TOP15 볼드
            if is_top15:
                for ci in [2, 3]:
                    ws.cell(row=row, column=ci).font = Font(bold=True, size=9)

        ws.freeze_panes = 'C2'

        # 컬러스케일 (매출·이동 점수)
        last = nrow + 1
        ws.conditional_formatting.add(f'E2:E{last}',
            ColorScaleRule(start_type='min', start_color='F4CCCC',
                           mid_type='num', mid_value=0, mid_color='FFFFFF',
                           end_type='max', end_color='B6D7A8'))
        ws.conditional_formatting.add(f'F2:F{last}',
            ColorScaleRule(start_type='min', start_color='F4CCCC',
                           mid_type='num', mid_value=0, mid_color='FFFFFF',
                           end_type='max', end_color='B6D7A8'))

    # 전체 시트 스타일
    ws_all = writer.sheets['전체']
    style_sheet(ws_all, len(out_df), 7, 'FFF2CC', '1F3864')

    # 사분면별 시트 스타일
    for q in QUAD_ORDER:
        ws_q = writer.sheets[q]
        n = len(out_df[out_df['사분면'] == q])
        style_sheet(ws_q, n, 7, QUAD_FILLS[q], QUAD_HEADER_FILLS[q])

print(f'[저장] {out_path}')
print(f'총 {len(merged)}개 행정동')
for q in QUAD_ORDER:
    cnt = (merged['사분면'] == q).sum()
    print(f'  {q}: {cnt}개 행정동')
