# -*- coding: utf-8 -*-
"""
2023~2024 분기별 20대 매출 QoQ 복합점수 TOP20 시각화 + 전체 순위 엑셀 출력
- 분석 분기: 2023Q1 ~ 2024Q4 (8분기), QoQ 베이스로 2022Q4 포함
- 3개 지표: QoQ성장률 / 20대매출규모 / 20대매출비중
  → 분기별 z-score → 지표별 8분기 평균 → 복합점수 = (z_규모 + z_비중 + z_성장률) / 3
출력: image/gen20_qoq_ranking.png
      data/gen20_qoq_ranking.xlsx
"""
import warnings
import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
import matplotlib.patches as mpatches
from scipy.stats import zscore as sp_zscore
from pathlib import Path
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule

warnings.filterwarnings('ignore')
plt.rcParams['font.family'] = 'Malgun Gothic'
plt.rcParams['axes.unicode_minus'] = False

ROOT = Path('.')
df = pd.read_csv(ROOT / 'data/main_data/분기별_원본피처.csv', encoding='utf-8-sig')

# ── 분기 설정 ─────────────────────────────────────────────────
# QoQ 계산을 위해 2022Q4(베이스) 포함
ALL_Q      = ['2022Q4','2023Q1','2023Q2','2023Q3','2023Q4',
              '2024Q1','2024Q2','2024Q3','2024Q4']
ANALYSIS_Q = ALL_Q[1:]   # 2023Q1 ~ 2024Q4 (8개)

Q_LABELS = {
    '2023Q1':'23Q1', '2023Q2':'23Q2', '2023Q3':'23Q3', '2023Q4':'23Q4',
    '2024Q1':'24Q1', '2024Q2':'24Q2', '2024Q3':'24Q3', '2024Q4':'24Q4',
}

# ── 분기별 집계 ───────────────────────────────────────────────
sub = df[df['연도분기'].isin(ALL_Q)]
qrt = (sub.groupby(['행정동', '연도분기'])[['연령대_20_매출_금액', '총_매출금액']]
          .sum())

piv_20  = qrt['연령대_20_매출_금액'].unstack('연도분기')
piv_tot = qrt['총_매출금액'].unstack('연도분기')

# 전 분기 데이터가 모두 있는 행정동만
piv_20  = piv_20.dropna(subset=ALL_Q)
piv_tot = piv_tot.reindex(piv_20.index).dropna(subset=ALL_Q)
common  = piv_20.index.intersection(piv_tot.index)
piv_20  = piv_20.loc[common]
piv_tot = piv_tot.loc[common]

# ── 분기별 3개 지표 계산 ──────────────────────────────────────
metrics = {}   # {분기: {'규모', '비중', '성장률'}}
for i, q in enumerate(ANALYSIS_Q):
    prev_q = ALL_Q[i]
    규모   = piv_20[q] / 1e8
    비중   = piv_20[q] / piv_tot[q] * 100
    성장률 = (piv_20[q] - piv_20[prev_q]) / piv_20[prev_q].abs() * 100
    metrics[q] = {'규모': 규모, '비중': 비중, '성장률': 성장률}

# ── 분기별 z-score → 지표별 8분기 평균 ──────────────────────
z_규모_list, z_비중_list, z_성장률_list = [], [], []

for q in ANALYSIS_Q:
    규모   = metrics[q]['규모']
    비중   = metrics[q]['비중']
    성장률 = metrics[q]['성장률']

    z_규모_list.append(pd.Series(sp_zscore(규모,   ddof=0), index=규모.index))
    z_비중_list.append(pd.Series(sp_zscore(비중,   ddof=0), index=비중.index))

    lo, hi = 성장률.quantile(0.005), 성장률.quantile(0.995)
    z_성장률_list.append(pd.Series(sp_zscore(성장률.clip(lo, hi), ddof=0), index=성장률.index))

z_규모_avg   = pd.concat(z_규모_list,   axis=1).mean(axis=1)
z_비중_avg   = pd.concat(z_비중_list,   axis=1).mean(axis=1)
z_성장률_avg = pd.concat(z_성장률_list, axis=1).mean(axis=1)

# ── 복합점수 ─────────────────────────────────────────────────
result = pd.DataFrame({
    'z_규모':   z_규모_avg,
    'z_비중':   z_비중_avg,
    'z_성장률': z_성장률_avg,
})
result['복합점수'] = (result['z_규모'] + result['z_비중'] + result['z_성장률']) / 3
result = result.sort_values('복합점수', ascending=False)

# 실제값 (2024Q4 기준으로 표기)
result['20대매출_24Q4(억)'] = piv_20['2024Q4'].reindex(result.index) / 1e8
result['20대비중_24Q4(%)']  = (piv_20['2024Q4'] / piv_tot['2024Q4'] * 100).reindex(result.index)

top20 = result.head(20).copy()

# ── 시각화 ───────────────────────────────────────────────────
fig, ax = plt.subplots(figsize=(16, 10))
fig.suptitle(
    "20대 매출 QoQ 복합점수 TOP 20  (2023Q1 ~ 2024Q4, 8분기)\n"
    "복합점수 = (z_규모_avg + z_비중_avg + z_성장률_avg) ÷ 3",
    fontsize=12, fontweight='bold', y=1.01
)

dongs  = top20.index[::-1]
scores = top20['복합점수'].values[::-1]
z_r    = top20['z_규모'].values[::-1]
z_b    = top20['z_비중'].values[::-1]
z_g    = top20['z_성장률'].values[::-1]

COLOR_HI = '#2E6FAD'
COLOR_LO = '#A8C8E8'
colors = [COLOR_HI if s >= 0.5 else COLOR_LO for s in scores]

ax.barh(range(len(dongs)), scores, color=colors,
        edgecolor='white', linewidth=0.6, height=0.72)

for i, (dong, sc, zr, zb, zg) in enumerate(zip(dongs, scores, z_r, z_b, z_g)):
    rank = len(dongs) - i
    ax.text(-0.03, i, f'{rank}위  {dong}',
            ha='right', va='center', fontsize=10, color='#333333')
    ax.text(sc + 0.015, i,
            f'{sc:.3f}   (규모 {zr:+.2f} / 비중 {zb:+.2f} / 성장 {zg:+.2f})',
            ha='left', va='center', fontsize=8.5, color='#555555')

ax.axvline(0,   color='#AAAAAA', lw=0.8)
ax.axvline(0.5, color='#CCCCCC', lw=0.8, ls='--')
ax.text(0.5, len(dongs) - 0.4, '복합=0.5', fontsize=7.5, color='#AAAAAA', ha='center')

ax.set_xlim(-0.2, scores.max() * 1.55)
ax.set_yticks([])
ax.set_xlabel('복합점수 (분기별 z-score 평균)', fontsize=11)
ax.spines[['top', 'right', 'left']].set_visible(False)
ax.grid(axis='x', alpha=0.2)
ax.legend(handles=[
    mpatches.Patch(color=COLOR_HI, label='복합점수 ≥ 0.5'),
    mpatches.Patch(color=COLOR_LO, label='복합점수 < 0.5'),
], fontsize=9, loc='lower right')

plt.tight_layout()
fig.savefig(ROOT / 'image/gen20_qoq_ranking.png', dpi=150, bbox_inches='tight')
plt.close(fig)
print('[저장] image/gen20_qoq_ranking.png')

# ── 엑셀 출력 ────────────────────────────────────────────────
Q_2023 = ['2023Q1','2023Q2','2023Q3','2023Q4']
Q_2024 = ['2024Q1','2024Q2','2024Q3','2024Q4']

# 연간 합계 및 파생 지표
sum23 = piv_20[Q_2023].sum(axis=1)
sum24 = piv_20[Q_2024].sum(axis=1)
result['23년_20대매출합계(억)']   = (sum23 / 1e8).reindex(result.index)
result['24년_20대매출합계(억)']   = (sum24 / 1e8).reindex(result.index)
result['23→24_성장률(%)']        = ((sum24 - sum23) / sum23.abs() * 100).reindex(result.index)

# 2023~2024 8분기 20대 비중 평균
비중_8q = pd.concat(
    [(piv_20[q] / piv_tot[q] * 100) for q in ANALYSIS_Q], axis=1
).mean(axis=1)
result['2023~24_20대비중평균(%)'] = 비중_8q.reindex(result.index)

# 분기별 원본 수치 추가
for q in ANALYSIS_Q:
    lbl = Q_LABELS[q]
    result[f'20대매출_{lbl}(억)'] = (piv_20[q] / 1e8).reindex(result.index)
    result[f'20대비중_{lbl}(%)']  = (piv_20[q] / piv_tot[q] * 100).reindex(result.index)
    result[f'QoQ성장률_{lbl}(%)'] = metrics[q]['성장률'].reindex(result.index)

result.insert(0, '순위', range(1, len(result) + 1))
result = result.reset_index()   # 행정동을 컬럼으로

# 컬럼 순서
core_cols   = ['순위', '행정동', '복합점수', 'z_규모', 'z_비중', 'z_성장률']
summary_cols = ['23년_20대매출합계(억)', '24년_20대매출합계(억)',
                '23→24_성장률(%)', '2023~24_20대비중평균(%)']
매출_cols   = [f'20대매출_{Q_LABELS[q]}(억)' for q in ANALYSIS_Q]
비중_cols   = [f'20대비중_{Q_LABELS[q]}(%)'  for q in ANALYSIS_Q]
성장률_cols = [f'QoQ성장률_{Q_LABELS[q]}(%)' for q in ANALYSIS_Q]
col_order   = core_cols + summary_cols + 매출_cols + 비중_cols + 성장률_cols
out_df = result[col_order]

out_path = ROOT / 'data/gen20_qoq_ranking.xlsx'
with pd.ExcelWriter(out_path, engine='openpyxl') as writer:
    out_df.to_excel(writer, index=False, sheet_name='QoQ복합순위')

    ws = writer.sheets['QoQ복합순위']
    nrow = len(out_df)
    ncol = len(col_order)
    last_row = nrow + 1

    # 헤더
    hdr_fill = PatternFill('solid', fgColor='1F3864')
    hdr_font = Font(bold=True, color='FFFFFF', size=9)
    for cell in ws[1]:
        cell.fill = hdr_fill
        cell.font = hdr_font
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    ws.row_dimensions[1].height = 32

    # 열 너비
    width_map = {1: 5, 2: 13, 3: 9, 4: 9, 5: 9, 6: 9}
    for i in range(1, ncol + 1):
        ws.column_dimensions[get_column_letter(i)].width = width_map.get(i, 10)

    # 숫자 포맷
    for row in range(2, nrow + 2):
        ws.row_dimensions[row].height = 15
        for col_idx, col_name in enumerate(col_order, 1):
            cell = ws.cell(row=row, column=col_idx)
            cell.alignment = Alignment(horizontal='center', vertical='center')
            if col_name == '복합점수':
                cell.number_format = '0.000'
            elif col_name.startswith('z_'):
                cell.number_format = '+0.00;-0.00'
            elif '억' in col_name:
                cell.number_format = '#,##0.0'
            elif '비중' in col_name or '성장률' in col_name:
                cell.number_format = '0.0"%"'
        ws.cell(row=row, column=2).alignment = Alignment(horizontal='left', vertical='center')

    # TOP 20 강조
    top20_fill = PatternFill('solid', fgColor='FFF2CC')
    for row in range(2, 22):
        for col in range(1, ncol + 1):
            ws.cell(row=row, column=col).fill = top20_fill
        for col in [1, 2, 3]:
            ws.cell(row=row, column=col).font = Font(bold=True, size=10)

    # 교차 배경 (21위~)
    alt_fill = PatternFill('solid', fgColor='F5F5F5')
    for row in range(22, nrow + 2):
        if row % 2 == 0:
            for col in range(1, ncol + 1):
                ws.cell(row=row, column=col).fill = alt_fill

    # 컬러스케일
    ws.conditional_formatting.add(f'C2:C{last_row}',
        ColorScaleRule(start_type='min', start_color='F8CBAD',
                       mid_type='percentile', mid_value=50, mid_color='FFFFFF',
                       end_type='max', end_color='2E75B6'))
    for col_letter in ['D', 'E', 'F']:
        ws.conditional_formatting.add(f'{col_letter}2:{col_letter}{last_row}',
            ColorScaleRule(start_type='min', start_color='F4CCCC',
                           mid_type='num', mid_value=0, mid_color='FFFFFF',
                           end_type='max', end_color='B6D7A8'))

    # 테두리
    thin = Side(style='thin', color='CCCCCC')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for row in ws.iter_rows(min_row=1, max_row=nrow+1, min_col=1, max_col=ncol):
        for cell in row:
            cell.border = border

    # 틀 고정
    ws.freeze_panes = 'C2'

print(f'[저장] {out_path}')
print(f'총 {nrow}개 행정동 / 컬럼 {ncol}개')
