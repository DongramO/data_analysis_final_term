# -*- coding: utf-8 -*-
"""
20대 주말 유입 '이동성장세 점수' 분석
3가지 지표 → Z-Score → 평균
  ① 장기 증가율   : 2023→2024 연간 합계 기준 YoY
  ② 추세 기울기   : 전체 분기(2020~2024) 선형회귀 기울기
  ③ 최근 모멘텀   : 최근 4분기(2024) vs 이전 4분기(2023) 분기 평균 비율
출력: image/gen20_migration_growth.png
      data/gen20_migration_growth.xlsx
"""
import warnings
import numpy as np
import pandas as pd
import matplotlib.pyplot as plt
import matplotlib.patches as mpatches
from scipy.stats import zscore as sp_zscore, linregress
from pathlib import Path
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule

warnings.filterwarnings('ignore')
plt.rcParams['font.family'] = 'Malgun Gothic'
plt.rcParams['axes.unicode_minus'] = False

ROOT     = Path('.')
MIG_PATH = ROOT / 'data/main_data/분기별_주말유입인구.csv'
MAP_PATH = ROOT / 'data/main_data/코드매핑_유입인구_매출.csv'
RF_PATH  = ROOT / 'data/main_data/분기별_원본피처.csv'

# ── 데이터 로드 및 행정동 매핑 ────────────────────────────────
_map = pd.read_csv(MAP_PATH, encoding='utf-8-sig')
_map.columns = ['mig_code', 'raw_code', 'gu', 'dong_map']

code2dong = (
    pd.read_csv(RF_PATH, encoding='utf-8-sig', usecols=['행정동_코드', '행정동'])
    .drop_duplicates()
    .rename(columns={'행정동_코드': 'raw_code', '행정동': 'dong'})
    .set_index('raw_code')['dong'].to_dict()
)

mig = pd.read_csv(MIG_PATH, encoding='utf-8-sig')
mig.columns = ['mig_code', 'yr', 'q', 'yrq', 'age', 'HE', 'WE', 'EE', 'mig_total']
mig['raw_code'] = mig['mig_code'].map(_map.set_index('mig_code')['raw_code'].to_dict())
mig = mig.dropna(subset=['raw_code'])
mig['raw_code'] = mig['raw_code'].astype(int)
mig['dong'] = mig['raw_code'].map(code2dong)
mig = mig.dropna(subset=['dong'])

# 20대만 (주말 유입 이미 필터됨)
mig20 = mig[mig['age'] == 20].copy()

# ── 행정동 × 분기 피벗 ───────────────────────────────────────
pivot = (mig20.groupby(['dong', 'yrq'])['mig_total']
              .sum()
              .unstack('yrq'))

Q_2023 = ['2023Q1','2023Q2','2023Q3','2023Q4']
Q_2024 = ['2024Q1','2024Q2','2024Q3','2024Q4']
ALL_Q  = Q_2023 + Q_2024   # 2023~2024 8분기만

# 분석 대상 8분기 모두 있는 행정동만
pivot = pivot.dropna(subset=ALL_Q)
print(f'분석 행정동: {len(pivot)}개')

# ── ① 장기 증가율 (2023→2024 연간 합계) ──────────────────────
sum_2023 = pivot[Q_2023].sum(axis=1)
sum_2024 = pivot[Q_2024].sum(axis=1)
long_growth = (sum_2024 - sum_2023) / sum_2023.abs() * 100

# ── ② 추세 기울기 (전체 20분기 선형회귀) ─────────────────────
X = np.arange(len(ALL_Q), dtype=float)
slopes = pivot[ALL_Q].apply(
    lambda row: linregress(X, row.values.astype(float)).slope, axis=1
)

# ── ③ 최근 모멘텀 (최근 4분기 vs 이전 4분기 평균) ────────────
recent_avg = pivot[Q_2024].mean(axis=1)
prev_avg   = pivot[Q_2023].mean(axis=1)
momentum   = (recent_avg - prev_avg) / prev_avg.abs() * 100

# ── Z-Score (0.5% 클리핑 후 표준화) ──────────────────────────
def clip_zscore(s):
    lo, hi = s.quantile(0.005), s.quantile(0.995)
    return pd.Series(sp_zscore(s.clip(lo, hi), ddof=0), index=s.index)

z_증가율 = clip_zscore(long_growth)
z_기울기 = clip_zscore(slopes)
z_모멘텀 = clip_zscore(momentum)

# ── 이동성장세 점수 ───────────────────────────────────────────
result = pd.DataFrame({
    'z_증가율': z_증가율,
    'z_기울기': z_기울기,
    'z_모멘텀': z_모멘텀,
    '장기증가율(%,23→24)':  long_growth,
    '추세기울기(명/분기)':   slopes,
    '최근모멘텀(%,24vs23)': momentum,
    '20대유입_2024(만명)':  sum_2024 / 1e4,
    '20대유입_2023(만명)':  sum_2023 / 1e4,
}, index=pivot.index)
result.index.name = '행정동'

result['이동성장세점수'] = (result['z_증가율'] + result['z_기울기'] + result['z_모멘텀']) / 3
result = result.sort_values('이동성장세점수', ascending=False)
result.insert(0, '순위', range(1, len(result) + 1))

top20 = result.head(20).copy()

# ── 시각화 ───────────────────────────────────────────────────
fig, ax = plt.subplots(figsize=(16, 10))
fig.suptitle(
    "20대 주말 유입 이동성장세 점수  TOP 20\n"
    "이동성장세점수 = (z_장기증가율 + z_추세기울기 + z_최근모멘텀) ÷ 3",
    fontsize=12, fontweight='bold', y=1.01
)

dongs  = top20.index[::-1]
scores = top20['이동성장세점수'].values[::-1]
z_g    = top20['z_증가율'].values[::-1]
z_s    = top20['z_기울기'].values[::-1]
z_m    = top20['z_모멘텀'].values[::-1]
v24    = top20['20대유입_2024(만명)'].values[::-1]

COLOR_HI = '#2E6FAD'
COLOR_LO = '#A8C8E8'
colors = [COLOR_HI if s >= 0.5 else COLOR_LO for s in scores]

ax.barh(range(len(dongs)), scores, color=colors,
        edgecolor='white', linewidth=0.6, height=0.72)

for i, (dong, sc, zg, zs, zm, v) in enumerate(
        zip(dongs, scores, z_g, z_s, z_m, v24)):
    rank = len(dongs) - i
    ax.text(-0.03, i, f'{rank}위  {dong}',
            ha='right', va='center', fontsize=10, color='#333333')
    ax.text(sc + 0.015, i,
            f'{sc:.3f}   (증가율 {zg:+.2f} / 기울기 {zs:+.2f} / 모멘텀 {zm:+.2f})  '
            f'[2024유입 {v:.1f}만]',
            ha='left', va='center', fontsize=8, color='#555555')

ax.axvline(0,   color='#AAAAAA', lw=0.8)
ax.axvline(0.5, color='#CCCCCC', lw=0.8, ls='--')
ax.text(0.5, len(dongs) - 0.4, '점수=0.5',
        fontsize=7.5, color='#AAAAAA', ha='center')

ax.set_xlim(-0.2, scores.max() * 1.65)
ax.set_yticks([])
ax.set_xlabel('이동성장세 점수 (z-score 평균)', fontsize=11)
ax.spines[['top', 'right', 'left']].set_visible(False)
ax.grid(axis='x', alpha=0.2)
ax.legend(handles=[
    mpatches.Patch(color=COLOR_HI, label='점수 ≥ 0.5'),
    mpatches.Patch(color=COLOR_LO, label='점수 < 0.5'),
], fontsize=9, loc='lower right')

plt.tight_layout()
img_path = ROOT / 'image/gen20_migration_growth.png'
fig.savefig(img_path, dpi=150, bbox_inches='tight')
plt.close(fig)
print(f'[저장] {img_path.name}')

# ── 엑셀 출력 ────────────────────────────────────────────────
# 분기별 원본 수치 추가
for q in ALL_Q:
    result[f'유입_{q}(만명)'] = (pivot[q] / 1e4).reindex(result.index)

out_df = result.reset_index()
col_order = (
    ['순위', '행정동', '이동성장세점수',
     'z_증가율', 'z_기울기', 'z_모멘텀',
     '장기증가율(%,23→24)', '추세기울기(명/분기)', '최근모멘텀(%,24vs23)',
     '20대유입_2024(만명)', '20대유입_2023(만명)']
    + [f'유입_{q}(만명)' for q in ALL_Q]
)
out_df = out_df[col_order]

out_path = ROOT / 'data/gen20_migration_growth_v2.xlsx'
with pd.ExcelWriter(out_path, engine='openpyxl') as writer:
    out_df.to_excel(writer, index=False, sheet_name='이동성장세순위')

    ws   = writer.sheets['이동성장세순위']
    nrow = len(out_df)
    ncol = len(col_order)
    last = nrow + 1

    # 헤더
    hf = PatternFill('solid', fgColor='1F3864')
    hfont = Font(bold=True, color='FFFFFF', size=9)
    for cell in ws[1]:
        cell.fill = hf; cell.font = hfont
        cell.alignment = Alignment(horizontal='center', vertical='center',
                                   wrap_text=True)
    ws.row_dimensions[1].height = 32

    # 열 너비
    w_map = {1:5, 2:13, 3:10, 4:9, 5:9, 6:9,
             7:14, 8:14, 9:14, 10:13, 11:13}
    for i in range(1, ncol + 1):
        ws.column_dimensions[get_column_letter(i)].width = w_map.get(i, 9)

    # 숫자 포맷 + 정렬
    for row in range(2, nrow + 2):
        ws.row_dimensions[row].height = 15
        for ci, cn in enumerate(col_order, 1):
            cell = ws.cell(row=row, column=ci)
            cell.alignment = Alignment(horizontal='center', vertical='center')
            if cn == '이동성장세점수':
                cell.number_format = '0.000'
            elif cn.startswith('z_'):
                cell.number_format = '+0.00;-0.00'
            elif '%' in cn or '증가율' in cn or '모멘텀' in cn:
                cell.number_format = '0.0"%"'
            elif '만명' in cn or '기울기' in cn:
                cell.number_format = '#,##0.0'
        ws.cell(row=row, column=2).alignment = Alignment(
            horizontal='left', vertical='center')

    # TOP 20 강조
    top20_fill = PatternFill('solid', fgColor='E8F4FD')
    for row in range(2, 22):
        for col in range(1, ncol + 1):
            ws.cell(row=row, column=col).fill = top20_fill
        for col in [1, 2, 3]:
            ws.cell(row=row, column=col).font = Font(bold=True, size=10)

    # 교차 배경 (21위~)
    alt = PatternFill('solid', fgColor='F5F5F5')
    for row in range(22, nrow + 2):
        if row % 2 == 0:
            for col in range(1, ncol + 1):
                ws.cell(row=row, column=col).fill = alt

    # 컬러스케일
    ws.conditional_formatting.add(f'C2:C{last}',
        ColorScaleRule(start_type='min', start_color='F8CBAD',
                       mid_type='percentile', mid_value=50, mid_color='FFFFFF',
                       end_type='max', end_color='2E75B6'))
    for cl in ['D', 'E', 'F']:
        ws.conditional_formatting.add(f'{cl}2:{cl}{last}',
            ColorScaleRule(start_type='min', start_color='F4CCCC',
                           mid_type='num', mid_value=0, mid_color='FFFFFF',
                           end_type='max', end_color='B6D7A8'))

    # 테두리
    thin   = Side(style='thin', color='CCCCCC')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for row in ws.iter_rows(min_row=1, max_row=nrow+1, min_col=1, max_col=ncol):
        for cell in row:
            cell.border = border

    ws.freeze_panes = 'C2'

print(f'[저장] {out_path}')
print(f'총 {nrow}개 행정동 / 컬럼 {ncol}개')
