# -*- coding: utf-8 -*-
"""
20대 주말 유입 이동성장세 점수
- ① z_증가율 : 2023→2024 연간 합계 단순 YoY (전체 성장 크기)
- ② z_기울기 : 2023Q1~2024Q4 8분기 선형회귀 기울기 (일관된 우상향 여부)
- ③ z_모멘텀 : QoQ 분기별(전분기 대비) 8분기 z-score 평균 (분기 가속도)
이동성장세점수 = (z_증가율 + z_기울기 + z_모멘텀) / 3
출력: image/gen20_migration_qoq.png
      data/gen20_migration_qoq.xlsx
"""
import warnings
import numpy as np
import pandas as pd
import matplotlib.pyplot as plt
import matplotlib.patches as mpatches
from scipy.stats import zscore as sp_zscore, linregress
from pathlib import Path
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule

warnings.filterwarnings('ignore')
plt.rcParams['font.family'] = 'Malgun Gothic'
plt.rcParams['axes.unicode_minus'] = False

ROOT     = Path('.')
MIG_PATH = ROOT / 'data/main_data/분기별_주말유입인구.csv'
MAP_PATH = ROOT / 'data/main_data/코드매핑_유입인구_매출.csv'
RF_PATH  = ROOT / 'data/main_data/분기별_원본피처.csv'

# ── 데이터 로드 및 행정동 매핑 ────────────────────────────────
_map = pd.read_csv(MAP_PATH, encoding='utf-8-sig')
_map.columns = ['mig_code', 'raw_code', 'gu', 'dong_map']

code2dong = (
    pd.read_csv(RF_PATH, encoding='utf-8-sig', usecols=['행정동_코드', '행정동'])
    .drop_duplicates()
    .rename(columns={'행정동_코드': 'raw_code', '행정동': 'dong'})
    .set_index('raw_code')['dong'].to_dict()
)

mig = pd.read_csv(MIG_PATH, encoding='utf-8-sig')
mig.columns = ['mig_code', 'yr', 'q', 'yrq', 'age', 'HE', 'WE', 'EE', 'mig_total']
mig['raw_code'] = mig['mig_code'].map(_map.set_index('mig_code')['raw_code'].to_dict())
mig = mig.dropna(subset=['raw_code'])
mig['raw_code'] = mig['raw_code'].astype(int)
mig['dong'] = mig['raw_code'].map(code2dong)
mig = mig.dropna(subset=['dong'])

# ── 분기 설정 ─────────────────────────────────────────────────
Q_2023   = ['2023Q1','2023Q2','2023Q3','2023Q4']
Q_2024   = ['2024Q1','2024Q2','2024Q3','2024Q4']
ANA_Q    = Q_2023 + Q_2024          # 분석 8분기
ALL_Q    = ['2022Q4'] + ANA_Q       # QoQ 베이스용 2022Q4 포함
Q_LABELS = {q: f"{q[2:4]}Q{q[-1]}" for q in ANA_Q}

# ── 20대 피벗 ────────────────────────────────────────────────
piv = (mig[mig['age'] == 20]
       .groupby(['dong', 'yrq'])['mig_total']
       .sum()
       .unstack('yrq'))

piv = piv.dropna(subset=ALL_Q)
print(f'분석 행정동: {len(piv)}개')

def clip_z(s):
    lo, hi = s.quantile(0.005), s.quantile(0.995)
    return pd.Series(sp_zscore(s.clip(lo, hi), ddof=0), index=s.index)

# ── ① z_증가율: 2023→2024 연간 합계 단순 YoY ─────────────────
sum_2023 = piv[Q_2023].sum(axis=1)
sum_2024 = piv[Q_2024].sum(axis=1)
증가율_raw = (sum_2024 - sum_2023) / sum_2023.abs() * 100
z_증가율 = clip_z(증가율_raw)

# ── ② z_기울기: 8분기 선형회귀 기울기 ───────────────────────
X = np.arange(len(ANA_Q), dtype=float)
slopes = piv[ANA_Q].apply(
    lambda row: linregress(X, row.values.astype(float)).slope, axis=1)
z_기울기 = clip_z(slopes)

# ── ③ z_모멘텀: QoQ 분기별 8분기 z-score 평균 ───────────────
qoq_raw = {}
z_모멘텀_list = []
for i, q in enumerate(ANA_Q):
    prev_q = ALL_Q[i]
    g = (piv[q] - piv[prev_q]) / piv[prev_q].abs() * 100
    qoq_raw[f'QoQ_{Q_LABELS[q]}(%)'] = g
    z_모멘텀_list.append(clip_z(g))

z_모멘텀 = pd.concat(z_모멘텀_list, axis=1).mean(axis=1)

# ── 이동성장세점수 ────────────────────────────────────────────
result = pd.DataFrame({
    'z_증가율': z_증가율,
    'z_기울기': z_기울기,
    'z_모멘텀': z_모멘텀,
}, index=piv.index)
result.index.name = '행정동'
result['이동성장세점수'] = (result['z_증가율'] + result['z_기울기'] + result['z_모멘텀']) / 3
result = result.sort_values('이동성장세점수', ascending=False)

result['20대유입_2024(만명)'] = (sum_2024 / 1e4).reindex(result.index)
result['20대유입_2023(만명)'] = (sum_2023 / 1e4).reindex(result.index)
result['연간증가율(%,23→24)'] = 증가율_raw.reindex(result.index)
result['추세기울기(명/분기)'] = slopes.reindex(result.index)

top20 = result.head(20).copy()

# ── 시각화 ───────────────────────────────────────────────────
fig, ax = plt.subplots(figsize=(16, 10))
fig.suptitle(
    "20대 주말 유입 이동성장세점수  TOP 20  (2023~2024)\n"
    "이동성장세점수 = (z_증가율 + z_기울기 + z_모멘텀) ÷ 3\n"
    "증가율: 연간YoY  |  기울기: 8분기 선형회귀  |  모멘텀: QoQ 8분기 평균",
    fontsize=11, fontweight='bold', y=1.02
)

dongs  = top20.index[::-1]
scores = top20['이동성장세점수'].values[::-1]
z_g    = top20['z_증가율'].values[::-1]
z_s    = top20['z_기울기'].values[::-1]
z_m    = top20['z_모멘텀'].values[::-1]

COLOR_HI = '#2E6FAD'
COLOR_LO = '#A8C8E8'
colors = [COLOR_HI if s >= 0.5 else COLOR_LO for s in scores]

ax.barh(range(len(dongs)), scores, color=colors,
        edgecolor='white', linewidth=0.6, height=0.72)

for i, (dong, sc, zg, zs, zm) in enumerate(zip(dongs, scores, z_g, z_s, z_m)):
    rank = len(dongs) - i
    ax.text(-0.03, i, f'{rank}위  {dong}',
            ha='right', va='center', fontsize=10, color='#333333')
    ax.text(sc + 0.015, i,
            f'{sc:.3f}   (증가율 {zg:+.2f} / 기울기 {zs:+.2f} / 모멘텀 {zm:+.2f})',
            ha='left', va='center', fontsize=8.5, color='#555555')

ax.axvline(0,   color='#AAAAAA', lw=0.8)
ax.axvline(0.5, color='#CCCCCC', lw=0.8, ls='--')
ax.text(0.5, len(dongs) - 0.4, '점수=0.5', fontsize=7.5, color='#AAAAAA', ha='center')
ax.set_xlim(-0.2, scores.max() * 1.55)
ax.set_yticks([])
ax.set_xlabel('이동성장세점수 (z-score 평균)', fontsize=11)
ax.spines[['top', 'right', 'left']].set_visible(False)
ax.grid(axis='x', alpha=0.2)
ax.legend(handles=[
    mpatches.Patch(color=COLOR_HI, label='점수 ≥ 0.5'),
    mpatches.Patch(color=COLOR_LO, label='점수 < 0.5'),
], fontsize=9, loc='lower right')

plt.tight_layout()
fig.savefig(ROOT / 'image/gen20_migration_qoq.png', dpi=150, bbox_inches='tight')
plt.close(fig)
print('[저장] image/gen20_migration_qoq.png')

# ── 엑셀 출력 ────────────────────────────────────────────────
for k, v in qoq_raw.items():
    result[k] = v.reindex(result.index)
for q in ANA_Q:
    result[f'유입_{Q_LABELS[q]}(만명)'] = (piv[q] / 1e4).reindex(result.index)

result.insert(0, '순위', range(1, len(result) + 1))
result = result.reset_index()

col_order = (
    ['순위', '행정동', '이동성장세점수', 'z_증가율', 'z_기울기', 'z_모멘텀',
     '20대유입_2024(만명)', '20대유입_2023(만명)', '연간증가율(%,23→24)', '추세기울기(명/분기)']
    + list(qoq_raw.keys())
    + [f'유입_{Q_LABELS[q]}(만명)' for q in ANA_Q]
)
out_df = result[col_order]

out_path = ROOT / 'data/gen20_migration_qoq.xlsx'
try:
    out_path.unlink(missing_ok=True)
except PermissionError:
    out_path = ROOT / 'data/gen20_migration_qoq_v3.xlsx'

with pd.ExcelWriter(out_path, engine='openpyxl') as writer:
    out_df.to_excel(writer, index=False, sheet_name='이동성장세_분기별')

    ws   = writer.sheets['이동성장세_분기별']
    nrow = len(out_df)
    ncol = len(col_order)
    last = nrow + 1

    hf = PatternFill('solid', fgColor='1F3864')
    hfont = Font(bold=True, color='FFFFFF', size=9)
    for cell in ws[1]:
        cell.fill = hf; cell.font = hfont
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    ws.row_dimensions[1].height = 32

    w_map = {1:5, 2:13, 3:11, 4:9, 5:9, 6:9, 7:13, 8:13, 9:14, 10:13}
    for i in range(1, ncol + 1):
        ws.column_dimensions[get_column_letter(i)].width = w_map.get(i, 11)

    for row in range(2, nrow + 2):
        ws.row_dimensions[row].height = 15
        for ci, cn in enumerate(col_order, 1):
            cell = ws.cell(row=row, column=ci)
            cell.alignment = Alignment(horizontal='center', vertical='center')
            if cn == '이동성장세점수':
                cell.number_format = '0.000'
            elif cn.startswith('z_'):
                cell.number_format = '+0.00;-0.00'
            elif '만명' in cn or '기울기' in cn:
                cell.number_format = '#,##0.0'
            elif '%' in cn:
                cell.number_format = '0.0"%"'
        ws.cell(row=row, column=2).alignment = Alignment(horizontal='left', vertical='center')

    top20_fill = PatternFill('solid', fgColor='E8F4FD')
    for row in range(2, 22):
        for col in range(1, ncol + 1):
            ws.cell(row=row, column=col).fill = top20_fill
        for col in [1, 2, 3]:
            ws.cell(row=row, column=col).font = Font(bold=True, size=10)

    alt = PatternFill('solid', fgColor='F5F5F5')
    for row in range(22, nrow + 2):
        if row % 2 == 0:
            for col in range(1, ncol + 1):
                ws.cell(row=row, column=col).fill = alt

    ws.conditional_formatting.add(f'C2:C{last}',
        ColorScaleRule(start_type='min', start_color='F8CBAD',
                       mid_type='percentile', mid_value=50, mid_color='FFFFFF',
                       end_type='max', end_color='2E75B6'))
    for cl in ['D', 'E', 'F']:
        ws.conditional_formatting.add(f'{cl}2:{cl}{last}',
            ColorScaleRule(start_type='min', start_color='F4CCCC',
                           mid_type='num', mid_value=0, mid_color='FFFFFF',
                           end_type='max', end_color='B6D7A8'))

    thin = Side(style='thin', color='CCCCCC')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for row in ws.iter_rows(min_row=1, max_row=nrow+1, min_col=1, max_col=ncol):
        for cell in row:
            cell.border = border

    ws.freeze_panes = 'C2'

print(f'[저장] {out_path.name}')
print(f'총 {nrow}개 행정동 / 컬럼 {ncol}개')
