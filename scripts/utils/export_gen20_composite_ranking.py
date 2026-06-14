# -*- coding: utf-8 -*-
"""
20대 매출 복합점수 전체 순위 엑셀 출력
성장률: 2020~2024 YoY 4개 z-score 평균 사용
출력: data/gen20_composite_ranking.xlsx
"""
import warnings
import pandas as pd
import numpy as np
from scipy.stats import zscore as sp_zscore
from pathlib import Path
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule

warnings.filterwarnings('ignore')

ROOT = Path('.')
df = pd.read_csv(ROOT / 'data/main_data/분기별_원본피처.csv', encoding='utf-8-sig')

# ── 연도별 집계 ───────────────────────────────────────────────
ann = (df.groupby(['행정동', '연도'])[['연령대_20_매출_금액', '총_매출금액']]
         .sum().reset_index())

years = [2020, 2021, 2022, 2023, 2024]
piv     = ann.pivot(index='행정동', columns='연도', values='연령대_20_매출_금액').dropna(subset=years)
piv_tot = ann.pivot(index='행정동', columns='연도', values='총_매출금액').reindex(piv.index)

# ── 지표 계산 ─────────────────────────────────────────────────
base = pd.DataFrame(index=piv.index)
base['20대매출(억,2024)']  = piv[2024] / 1e8
base['총매출(억,2024)']    = piv_tot[2024] / 1e8
base['20대비중(%,2024)']   = base['20대매출(억,2024)'] / base['총매출(억,2024)'] * 100

YOY_PAIRS = [(2020,2021), (2021,2022), (2022,2023), (2023,2024)]
for y0, y1 in YOY_PAIRS:
    base[f'YoY성장률_{y0}→{y1}(%)'] = (piv[y1] - piv[y0]) / piv[y0].abs() * 100

# 4년 전체 성장률 및 CAGR
base['전체성장률_2020→2024(%)'] = (piv[2024] - piv[2020]) / piv[2020].abs() * 100
base['CAGR_2020→2024(%)'] = ((piv[2024] / piv[2020]).abs() ** (1/4) - 1) * 100

# ── z-score 변환 ─────────────────────────────────────────────
base['z_규모'] = sp_zscore(base['20대매출(억,2024)'], ddof=0)
base['z_비중'] = sp_zscore(base['20대비중(%,2024)'],  ddof=0)

yoy_z_cols = []
for y0, y1 in YOY_PAIRS:
    raw = base[f'YoY성장률_{y0}→{y1}(%)']
    lo, hi = raw.quantile(0.005), raw.quantile(0.995)
    z_col = f'z_성장률_{y0}→{y1}'
    base[z_col] = sp_zscore(raw.clip(lo, hi), ddof=0)
    yoy_z_cols.append(z_col)

base['z_성장률_avg(4YoY)'] = base[yoy_z_cols].mean(axis=1)

# ── 복합점수 및 순위 ──────────────────────────────────────────
base['복합점수'] = (base['z_규모'] + base['z_비중'] + base['z_성장률_avg(4YoY)']) / 3
base = base.sort_values('복합점수', ascending=False)
base.insert(0, '순위', range(1, len(base) + 1))
base = base.reset_index()

# ── 컬럼 순서 정리 ────────────────────────────────────────────
col_order = (
    ['순위', '행정동', '복합점수',
     'z_규모', 'z_비중', 'z_성장률_avg(4YoY)']
    + yoy_z_cols
    + ['20대매출(억,2024)', '20대비중(%,2024)', '총매출(억,2024)']
    + [f'YoY성장률_{y0}→{y1}(%)' for y0,y1 in YOY_PAIRS]
    + ['전체성장률_2020→2024(%)', 'CAGR_2020→2024(%)']
)
out_df = base[col_order]

# ── 엑셀 저장 ────────────────────────────────────────────────
out_path = ROOT / 'data/gen20_composite_ranking_v3.xlsx'

with pd.ExcelWriter(out_path, engine='openpyxl') as writer:
    out_df.to_excel(writer, index=False, sheet_name='전체순위')

    wb = writer.book
    ws = writer.sheets['전체순위']
    nrow = len(out_df)
    ncol = len(col_order)

    # ── 헤더 스타일 ──────────────────────────────────────────
    header_fill = PatternFill('solid', fgColor='1F3864')
    header_font = Font(bold=True, color='FFFFFF', size=9)
    center = Alignment(horizontal='center', vertical='center', wrap_text=True)

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center
    ws.row_dimensions[1].height = 30

    # ── 열 너비 ──────────────────────────────────────────────
    width_map = {1: 5, 2: 13, 3: 9}  # 순위, 행정동, 복합점수
    default_w = 11
    for i in range(1, ncol + 1):
        w = width_map.get(i, default_w)
        ws.column_dimensions[get_column_letter(i)].width = w

    # ── 숫자 포맷 ────────────────────────────────────────────
    fmt3 = '0.000'
    fmt2p = '+0.00;-0.00'
    fmt1  = '#,##0.0'
    fmt1p = '0.0"%"'

    for row in range(2, nrow + 2):
        ws.row_dimensions[row].height = 15
        for col_idx, col_name in enumerate(col_order, 1):
            cell = ws.cell(row=row, column=col_idx)
            cell.alignment = Alignment(horizontal='center', vertical='center')
            if col_name == '복합점수':
                cell.number_format = fmt3
            elif col_name.startswith('z_'):
                cell.number_format = fmt2p
            elif '매출(억' in col_name:
                cell.number_format = fmt1
            elif '비중' in col_name or '성장률' in col_name or 'CAGR' in col_name:
                cell.number_format = fmt1p
        ws.cell(row=row, column=2).alignment = Alignment(horizontal='left', vertical='center')

    # ── TOP 20 강조 ──────────────────────────────────────────
    top20_fill = PatternFill('solid', fgColor='FFF2CC')
    for row in range(2, 22):
        for col in range(1, ncol + 1):
            ws.cell(row=row, column=col).fill = top20_fill
        for col in [1, 2, 3]:
            ws.cell(row=row, column=col).font = Font(bold=True, size=10)

    # ── 교차 배경 (21위~) ────────────────────────────────────
    alt_fill = PatternFill('solid', fgColor='F5F5F5')
    for row in range(22, nrow + 2):
        if row % 2 == 0:
            for col in range(1, ncol + 1):
                ws.cell(row=row, column=col).fill = alt_fill

    # ── 컬러스케일 ───────────────────────────────────────────
    last_row = nrow + 1
    ws.conditional_formatting.add(f'C2:C{last_row}',
        ColorScaleRule(start_type='min', start_color='F8CBAD',
                       mid_type='percentile', mid_value=50, mid_color='FFFFFF',
                       end_type='max', end_color='2E75B6'))
    for col_letter in ['D', 'E', 'F']:
        ws.conditional_formatting.add(f'{col_letter}2:{col_letter}{last_row}',
            ColorScaleRule(start_type='min', start_color='F4CCCC',
                           mid_type='num', mid_value=0, mid_color='FFFFFF',
                           end_type='max', end_color='B6D7A8'))

    # ── 테두리 ───────────────────────────────────────────────
    thin = Side(style='thin', color='CCCCCC')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for row in ws.iter_rows(min_row=1, max_row=nrow+1, min_col=1, max_col=ncol):
        for cell in row:
            cell.border = border

    # ── 틀 고정 (행정동·순위 고정) ───────────────────────────
    ws.freeze_panes = 'C2'

print(f'[저장] {out_path}')
print(f'총 {nrow}개 행정동 / 컬럼 {ncol}개')
